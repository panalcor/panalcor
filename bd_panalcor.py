"""
PANALCOR S.L. — Base de datos SQLite paralela (panalcor.db)
============================================================
Sincroniza los datos existentes (Excel + Firebase) a una base de datos
SQLite bien estructurada, SIN tocar la forma de trabajo actual.

Uso directo:    python bd_panalcor.py          → sincronización completa
Desde servidor: importado por servidor.py      → sync periódico + API consultas

Requisitos: Python 3.7+ (sqlite3 y urllib incluidos).
            openpyxl solo hace falta para importar los Excel (si no está,
            se sincroniza únicamente Firebase).
"""

import json
import sqlite3
import ssl
import datetime
import threading
import urllib.request
from pathlib import Path

BASE_DIR = Path(__file__).parent
DB_FILE  = BASE_DIR / 'panalcor.db'

FIREBASE_URL = 'https://panalcor-mantenimiento-default-rtdb.europe-west1.firebasedatabase.app'

# Contexto SSL: usar el almacén de certificados de Windows si truststore
# está disponible (necesario cuando un antivirus inspecciona el HTTPS).
_SSL_CTX = None
try:
    import truststore
    _SSL_CTX = truststore.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
except ImportError:
    try:
        _SSL_CTX = ssl.create_default_context()
    except Exception:
        _SSL_CTX = None

_lock = threading.Lock()


# ════════════════════════════════════════════════════════════════
# CONEXIÓN Y ESQUEMA
# ════════════════════════════════════════════════════════════════
def conectar():
    con = sqlite3.connect(DB_FILE, timeout=15)
    con.row_factory = sqlite3.Row
    con.execute('PRAGMA journal_mode=WAL')
    return con


SCHEMA = """
CREATE TABLE IF NOT EXISTS maquinas (
    id            TEXT PRIMARY KEY,
    nombre        TEXT, tipo TEXT, fabricante TEXT, area TEXT,
    criticidad    TEXT, modelo TEXT, n_serie TEXT, estado TEXT,
    notas         TEXT, uso TEXT, mantenimiento TEXT, repuestos TEXT,
    advertencias  TEXT,
    origen        TEXT,            -- 'excel' | 'firebase'
    actualizado   TEXT
);

CREATE TABLE IF NOT EXISTS mecanicos (
    id        TEXT PRIMARY KEY,
    nombre    TEXT, turno TEXT, especialidad TEXT, especialidad2 TEXT,
    nivel     TEXT, estado TEXT, telefono TEXT, email TEXT, notas TEXT,
    origen    TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS repuestos (
    referencia        TEXT PRIMARY KEY,
    ubicacion         TEXT, descripcion TEXT, familia TEXT,
    stock             REAL, stock_minimo REAL,
    fecha_ult_entrada TEXT, fecha_ult_salida TEXT,
    precio            REAL, proveedor TEXT, notas TEXT,
    actualizado       TEXT
);

CREATE TABLE IF NOT EXISTS averias (
    id           TEXT PRIMARY KEY,
    fecha        TEXT, hora_inicio TEXT, hora_fin TEXT, tiempo_min REAL,
    maq_id       TEXT, maq_nombre TEXT, maq_tipo TEXT, maq_area TEXT,
    tipo         TEXT, causa TEXT, gravedad TEXT, criticidad TEXT,
    parte        TEXT, descripcion TEXT, accion TEXT, observaciones TEXT,
    estado       TEXT, turno TEXT, ts INTEGER,
    json         TEXT,              -- registro original completo
    actualizado  TEXT
);
CREATE INDEX IF NOT EXISTS idx_averias_fecha ON averias(fecha);
CREATE INDEX IF NOT EXISTS idx_averias_maq   ON averias(maq_id);

CREATE TABLE IF NOT EXISTS averia_materiales (
    averia_id   TEXT, referencia TEXT, descripcion TEXT, cantidad REAL,
    PRIMARY KEY (averia_id, referencia, descripcion)
);
CREATE INDEX IF NOT EXISTS idx_avmat_ref ON averia_materiales(referencia);

CREATE TABLE IF NOT EXISTS averia_mecanicos (
    averia_id TEXT, mecanico TEXT,
    PRIMARY KEY (averia_id, mecanico)
);

CREATE TABLE IF NOT EXISTS tareas_diarias (
    id TEXT PRIMARY KEY,
    area TEXT, maquina TEXT, tarea TEXT, detalle TEXT,
    prioridad TEXT, turno TEXT, orden INTEGER, json TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS preventivo_catalogo (
    id TEXT PRIMARY KEY,
    maquina TEXT, subtipo TEXT, parte TEXT, tarea TEXT,
    frecuencia TEXT, instancias TEXT, json TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS preventivo_asignaciones (
    clave TEXT PRIMARY KEY,
    catalog_id TEXT, maquina TEXT, instancia TEXT, subtipo TEXT,
    tarea TEXT, frecuencia TEXT, mes TEXT, mecanico TEXT,
    hecho INTEGER, notas TEXT, json TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS mant_anual (
    id TEXT PRIMARY KEY,
    mes TEXT, tarea TEXT, descripcion TEXT, maquina TEXT,
    responsable TEXT, periodicidad TEXT, fecha_inicio TEXT,
    horas_est REAL, completado TEXT, fecha_realizado TEXT,
    turno TEXT, mecanico TEXT, observaciones TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS ausencias (
    clave TEXT PRIMARY KEY,
    mec_id TEXT, tipo TEXT, desde TEXT, hasta TEXT, nota TEXT,
    ts INTEGER, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS pedidos (
    clave TEXT PRIMARY KEY,
    referencia TEXT, descripcion TEXT, cantidad REAL, mecanico TEXT,
    estado TEXT, motivo TEXT, nota_compras TEXT,
    ts INTEGER, ts_compra INTEGER, json TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS recordatorios (
    clave TEXT PRIMARY KEY,
    fecha TEXT, hora TEXT, descripcion TEXT, mecanico TEXT,
    prioridad TEXT, estado TEXT, creado_en TEXT, asignado_en TEXT,
    json TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS estado_maquinas (
    maq_id TEXT PRIMARY KEY,
    estado TEXT, nota TEXT, ts INTEGER, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS frio_equipos (
    clave TEXT PRIMARY KEY,
    nombre TEXT, tipo TEXT, ubicacion TEXT, estado TEXT,
    marca TEXT, modelo TEXT, refrigerante TEXT,
    t_min TEXT, t_max TEXT, json TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS frio_averias (
    clave TEXT PRIMARY KEY,
    equipo_key TEXT, fecha_hora TEXT, descripcion TEXT, causa TEXT,
    solucion TEXT, mecanico TEXT, estado TEXT, ts INTEGER,
    json TEXT, actualizado TEXT
);

CREATE TABLE IF NOT EXISTS frio_lecturas (
    clave TEXT PRIMARY KEY,
    eq_id TEXT, eq_nombre TEXT, fecha TEXT, hora TEXT, obs TEXT,
    json TEXT, actualizado TEXT
);

-- Nodos de Firebase sin tabla propia: se guardan completos igualmente
CREATE TABLE IF NOT EXISTS fb_otros (
    nodo TEXT, clave TEXT, json TEXT, actualizado TEXT,
    PRIMARY KEY (nodo, clave)
);

CREATE TABLE IF NOT EXISTS sync_log (
    ts TEXT, origen TEXT, detalle TEXT
);
"""


def crear_esquema():
    with conectar() as con:
        con.executescript(SCHEMA)


def _ahora():
    return datetime.datetime.now().isoformat(timespec='seconds')


def _log_sync(con, origen, detalle):
    con.execute('INSERT INTO sync_log VALUES (?,?,?)', (_ahora(), origen, detalle))


# ════════════════════════════════════════════════════════════════
# FIREBASE (REST, solo lectura)
# ════════════════════════════════════════════════════════════════
def _fb_get(nodo):
    """Lee un nodo de Firebase. Devuelve None si no hay permiso o falla."""
    url = f'{FIREBASE_URL}/{nodo}.json'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'panalcor-sync'})
        with urllib.request.urlopen(req, timeout=30, context=_SSL_CTX) as r:
            return json.load(r)
    except Exception:
        return None


def _upsert(con, tabla, columnas, valores):
    marcas = ','.join('?' * len(columnas))
    cols   = ','.join(columnas)
    con.execute(
        f'INSERT INTO {tabla} ({cols}) VALUES ({marcas}) '
        f'ON CONFLICT DO UPDATE SET ' +
        ','.join(f'{c}=excluded.{c}' for c in columnas[1:]),
        valores)


def _sync_averias(con, datos, origen='firebase'):
    n = 0
    for clave, a in (datos or {}).items():
        if not isinstance(a, dict):
            continue
        aid = a.get('id') or clave
        _upsert(con, 'averias',
            ['id','fecha','hora_inicio','hora_fin','tiempo_min','maq_id','maq_nombre',
             'maq_tipo','maq_area','tipo','causa','gravedad','criticidad','parte',
             'descripcion','accion','observaciones','estado','turno','ts','json','actualizado'],
            [aid, a.get('fecha'), a.get('horaInicio'), a.get('horaFin'), a.get('tiempoMin'),
             a.get('maqId'), a.get('maqNombre'), a.get('maqTipo'), a.get('maqArea'),
             a.get('tipo'), a.get('causa'), a.get('gravedad'), a.get('criticidad'),
             a.get('parte'), a.get('descripcion'), a.get('accion'), a.get('observaciones'),
             a.get('estado'), a.get('turno'), a.get('ts'),
             json.dumps(a, ensure_ascii=False), _ahora()])
        # Materiales usados (tabla hija)
        con.execute('DELETE FROM averia_materiales WHERE averia_id=?', (aid,))
        for m in a.get('materiales') or []:
            if isinstance(m, dict):
                con.execute('INSERT OR REPLACE INTO averia_materiales VALUES (?,?,?,?)',
                            (aid, str(m.get('ref') or ''), m.get('desc'), m.get('qty')))
        # Mecánicos que intervinieron (tabla hija)
        con.execute('DELETE FROM averia_mecanicos WHERE averia_id=?', (aid,))
        for mec in a.get('mecanicos') or []:
            con.execute('INSERT OR REPLACE INTO averia_mecanicos VALUES (?,?)', (aid, str(mec)))
        n += 1
    return n


def guardar_averia(averia):
    """Guarda UNA avería (la llama servidor.py al recibirla del móvil)."""
    with _lock, conectar() as con:
        crear_esquema()
        _sync_averias(con, {averia.get('id', 'sin-id'): averia}, origen='movil')


def sync_firebase():
    """Sincroniza todos los nodos legibles de Firebase a SQLite."""
    resumen = []
    with _lock, conectar() as con:
        crear_esquema()

        d = _fb_get('averias')
        if d: resumen.append(f"averias:{_sync_averias(con, d)}")

        d = _fb_get('maquinas_extra')
        for clave, m in (d or {}).items():
            if not isinstance(m, dict): continue
            _upsert(con, 'maquinas',
                ['id','nombre','tipo','fabricante','area','origen','actualizado'],
                [m.get('id') or clave, m.get('n'), m.get('t'), m.get('f'), m.get('a'),
                 'firebase', _ahora()])
        if d: resumen.append(f"maquinas_extra:{len(d)}")

        d = _fb_get('config/personal')
        regs = d.items() if isinstance(d, dict) else enumerate(d or [])
        n = 0
        for clave, p in regs:
            if not isinstance(p, dict): continue
            _upsert(con, 'mecanicos',
                ['id','nombre','turno','especialidad','estado','notas','origen','actualizado'],
                [p.get('id') or str(clave), p.get('nombre'), p.get('turno'),
                 p.get('espec'), p.get('estado'), p.get('notas'), 'firebase', _ahora()])
            n += 1
        if n: resumen.append(f"personal:{n}")

        d = _fb_get('tareas_diarias')
        for clave, t in (d or {}).items():
            if not isinstance(t, dict): continue
            _upsert(con, 'tareas_diarias',
                ['id','area','maquina','tarea','detalle','prioridad','turno','orden','json','actualizado'],
                [t.get('id') or clave, t.get('area'), t.get('maquina'), t.get('tarea'),
                 t.get('detalle'), t.get('prioridad'), t.get('turno'), t.get('orden'),
                 json.dumps(t, ensure_ascii=False), _ahora()])
        if d: resumen.append(f"tareas_diarias:{len(d)}")

        d = _fb_get('catalogo')
        for clave, c in (d or {}).items():
            if not isinstance(c, dict): continue
            _upsert(con, 'preventivo_catalogo',
                ['id','maquina','subtipo','parte','tarea','frecuencia','instancias','json','actualizado'],
                [c.get('id') or clave, c.get('maquina'), c.get('subtipo'), c.get('parte'),
                 c.get('tarea'), c.get('frecuencia'),
                 json.dumps(c.get('instancias') or [], ensure_ascii=False),
                 json.dumps(c, ensure_ascii=False), _ahora()])
        if d: resumen.append(f"catalogo:{len(d)}")

        d = _fb_get('asignaciones')
        n = 0
        for mes_o_grupo, grupo in (d or {}).items():
            if not isinstance(grupo, dict): continue
            for clave, a in grupo.items():
                if not isinstance(a, dict): continue
                _upsert(con, 'preventivo_asignaciones',
                    ['clave','catalog_id','maquina','instancia','subtipo','tarea','frecuencia',
                     'mes','mecanico','hecho','notas','json','actualizado'],
                    [f'{mes_o_grupo}/{clave}', a.get('catalogId'), a.get('maquina'),
                     a.get('instancia'), a.get('subtipo'), a.get('tarea'), a.get('frecuencia'),
                     a.get('mes'), a.get('mecanico'), 1 if a.get('hecho') else 0,
                     a.get('notas'), json.dumps(a, ensure_ascii=False), _ahora()])
                n += 1
        if n: resumen.append(f"asignaciones:{n}")

        d = _fb_get('turnos')
        n = 0
        for grupo, contenido in (d or {}).items():
            if grupo == 'ausencias' and isinstance(contenido, dict):
                for clave, au in contenido.items():
                    if not isinstance(au, dict): continue
                    _upsert(con, 'ausencias',
                        ['clave','mec_id','tipo','desde','hasta','nota','ts','actualizado'],
                        [clave, au.get('mecId'), au.get('tipo'), au.get('desde'),
                         au.get('hasta'), au.get('nota'), au.get('ts'), _ahora()])
                    n += 1
            else:
                _upsert(con, 'fb_otros', ['nodo','clave','json','actualizado'],
                        ['turnos', grupo, json.dumps(contenido, ensure_ascii=False), _ahora()])
        if n: resumen.append(f"ausencias:{n}")

        d = _fb_get('pedidos')
        for clave, p in (d or {}).items():
            if not isinstance(p, dict): continue
            _upsert(con, 'pedidos',
                ['clave','referencia','descripcion','cantidad','mecanico','estado','motivo',
                 'nota_compras','ts','ts_compra','json','actualizado'],
                [clave, p.get('ref'), p.get('desc'), p.get('cantidad'), p.get('mecanico'),
                 p.get('estado'), p.get('motivo'), p.get('notaCompras'), p.get('ts'),
                 p.get('tsCompra'), json.dumps(p, ensure_ascii=False), _ahora()])
        if d: resumen.append(f"pedidos:{len(d)}")

        d = _fb_get('recordatorios')
        for clave, r in (d or {}).items():
            if not isinstance(r, dict): continue
            _upsert(con, 'recordatorios',
                ['clave','fecha','hora','descripcion','mecanico','prioridad','estado',
                 'creado_en','asignado_en','json','actualizado'],
                [clave, r.get('fecha'), r.get('hora'), r.get('descripcion'), r.get('mecanico'),
                 r.get('prioridad'), r.get('estado'), r.get('creadoEn'), r.get('asignadoEn'),
                 json.dumps(r, ensure_ascii=False), _ahora()])
        if d: resumen.append(f"recordatorios:{len(d)}")

        d = _fb_get('estado_maquinas')
        for clave, e in (d or {}).items():
            if not isinstance(e, dict): continue
            _upsert(con, 'estado_maquinas',
                ['maq_id','estado','nota','ts','actualizado'],
                [clave, e.get('estado'), e.get('nota'), e.get('ts'), _ahora()])
        if d: resumen.append(f"estado_maquinas:{len(d)}")

        d = _fb_get('frio/equipos')
        for clave, e in (d or {}).items():
            if not isinstance(e, dict): continue
            _upsert(con, 'frio_equipos',
                ['clave','nombre','tipo','ubicacion','estado','marca','modelo','refrigerante',
                 't_min','t_max','json','actualizado'],
                [clave, e.get('nombre'), e.get('tipo'), e.get('ubicacion'), e.get('estado'),
                 e.get('marca'), e.get('modelo'), e.get('refrigerante'),
                 str(e.get('tMin') or ''), str(e.get('tMax') or ''),
                 json.dumps(e, ensure_ascii=False), _ahora()])
        if d: resumen.append(f"frio_equipos:{len(d)}")

        d = _fb_get('frio/averias')
        for clave, a in (d or {}).items():
            if not isinstance(a, dict): continue
            _upsert(con, 'frio_averias',
                ['clave','equipo_key','fecha_hora','descripcion','causa','solucion',
                 'mecanico','estado','ts','json','actualizado'],
                [clave, a.get('equipoKey'), a.get('fechaHora'), a.get('descripcion'),
                 a.get('causa'), a.get('solucion'), a.get('mecanico'), a.get('estado'),
                 a.get('ts'), json.dumps(a, ensure_ascii=False), _ahora()])
        if d: resumen.append(f"frio_averias:{len(d)}")

        d = _fb_get('frio/lecturas')
        for clave, l in (d or {}).items():
            if not isinstance(l, dict): continue
            _upsert(con, 'frio_lecturas',
                ['clave','eq_id','eq_nombre','fecha','hora','obs','json','actualizado'],
                [l.get('id') or clave, l.get('eqId'), l.get('eqNombre'), l.get('fecha'),
                 l.get('hora'), l.get('obs'), json.dumps(l, ensure_ascii=False), _ahora()])
        if d: resumen.append(f"frio_lecturas:{len(d)}")

        d = _fb_get('repuestos_catalogo')
        for clave, r in (d or {}).items():
            if not isinstance(r, dict): continue
            ref = str(r.get('ref') or clave)
            # Enriquecer la fila del almacén (si la referencia no existe en
            # repuestos.xlsx, el catálogo no añade filas nuevas)
            con.execute("""
                UPDATE repuestos SET precio=?, proveedor=?, notas=?
                WHERE referencia=?""",
                (r.get('precio'), r.get('prov'), r.get('notas'), ref))
        if d: resumen.append(f"repuestos_catalogo:{len(d)}")

        # Nodos sin estructura propia: guardar tal cual para no perder nada
        for nodo in ['registros_turno','cursos','integrales','sugerencias','proyectos',
                     'avisos_encargados','frio/temperaturas','frio/preventivo',
                     'frio/checklist_custom','config/ultima_realizacion']:
            d = _fb_get(nodo)
            if isinstance(d, dict) and d:
                for clave, v in d.items():
                    _upsert(con, 'fb_otros', ['nodo','clave','json','actualizado'],
                            [nodo, str(clave), json.dumps(v, ensure_ascii=False), _ahora()])
                resumen.append(f"{nodo}:{len(d)}")

        _log_sync(con, 'firebase', ', '.join(resumen) or 'sin datos')
    return resumen


# ════════════════════════════════════════════════════════════════
# EXCEL (importación inicial / refresco)
# ════════════════════════════════════════════════════════════════
def _celda(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    return str(v).strip() if isinstance(v, str) else v


def sync_excel():
    """Importa los Excel del proyecto. Devuelve resumen de filas."""
    try:
        import openpyxl
    except ImportError:
        return ['openpyxl no instalado — Excel omitidos']

    resumen = []
    with _lock, conectar() as con:
        crear_esquema()

        # ── Máquinas ─────────────────────────────────────────────
        f = BASE_DIR / 'BD_Maquinaria_PANALCOR_2026-05-24.xlsx'
        if f.exists():
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb['Maquinaria']
            n = 0
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if not fila or not fila[0]: continue
                v = [_celda(x) for x in fila[:14]]
                v += [None] * (14 - len(v))
                _upsert(con, 'maquinas',
                    ['id','nombre','tipo','fabricante','area','criticidad','modelo','n_serie',
                     'estado','notas','uso','mantenimiento','repuestos','advertencias',
                     'origen','actualizado'],
                    v + ['excel', _ahora()])
                n += 1
            wb.close()
            resumen.append(f'maquinas:{n}')

        # ── Mecánicos (sin contraseñas) ──────────────────────────
        f = BASE_DIR / 'BD_Mecanicos.xlsx'
        if f.exists():
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            n = 0
            for fila in ws.iter_rows(min_row=4, values_only=True):
                if not fila or not fila[0]: continue
                v = [_celda(x) for x in fila[:9]]   # se excluye col 10 (Contraseña App)
                v += [None] * (9 - len(v))
                _upsert(con, 'mecanicos',
                    ['id','nombre','turno','especialidad','especialidad2','nivel','estado',
                     'telefono','email','origen','actualizado'],
                    v + ['excel', _ahora()])
                n += 1
            wb.close()
            resumen.append(f'mecanicos:{n}')

        # ── Repuestos (almacén) ──────────────────────────────────
        f = BASE_DIR / 'repuestos.xlsx'
        if f.exists():
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb['Tabla1']
            n = 0
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if not fila or fila[1] is None: continue
                v = [_celda(x) for x in fila[:8]]
                v += [None] * (8 - len(v))
                _upsert(con, 'repuestos',
                    ['referencia','ubicacion','descripcion','familia','stock','stock_minimo',
                     'fecha_ult_entrada','fecha_ult_salida','actualizado'],
                    [str(v[1]), v[0], v[2], v[3], v[4], v[5], v[6], v[7], _ahora()])
                n += 1
            wb.close()
            resumen.append(f'repuestos:{n}')

        # ── Mantenimiento anual ──────────────────────────────────
        f = BASE_DIR / 'anual.xlsx'
        if f.exists():
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            n = 0
            for fila in ws.iter_rows(min_row=4, values_only=True):
                if not fila or not fila[0]: continue
                v = [_celda(x) for x in fila[:14]]
                v += [None] * (14 - len(v))
                _upsert(con, 'mant_anual',
                    ['id','mes','tarea','descripcion','maquina','responsable','periodicidad',
                     'fecha_inicio','horas_est','completado','fecha_realizado','turno',
                     'mecanico','observaciones','actualizado'],
                    [str(v[0])] + v[1:] + [_ahora()])
                n += 1
            wb.close()
            resumen.append(f'mant_anual:{n}')

        # ── Tareas diarias (Excel; Firebase tareas_diarias manda) ─
        f = BASE_DIR / 'diarios.xlsx'
        if f.exists():
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            n = 0
            for fila in ws.iter_rows(min_row=4, values_only=True):
                if not fila or not fila[0]: continue
                v = [_celda(x) for x in fila[:12]]
                v += [None] * (12 - len(v))
                _upsert(con, 'tareas_diarias',
                    ['id','tarea','maquina','area','detalle','turno','prioridad','json','actualizado'],
                    [str(v[0]), v[1], v[2], v[3], v[4], v[5], v[6],
                     json.dumps({'origen':'diarios.xlsx'}, ensure_ascii=False), _ahora()])
                n += 1
            wb.close()
            resumen.append(f'diarios:{n}')

        _log_sync(con, 'excel', ', '.join(resumen) or 'sin archivos')
    return resumen


def sync_completo():
    """Excel + Firebase. Devuelve resumen combinado."""
    r1 = sync_excel()
    r2 = sync_firebase()
    return {'excel': r1, 'firebase': r2}


def sync_periodico(intervalo=900):
    """Hilo en segundo plano: sincroniza Firebase cada `intervalo` segundos."""
    def _bucle():
        import time
        while True:
            try:
                sync_firebase()
            except Exception:
                pass
            time.sleep(intervalo)
    t = threading.Thread(target=_bucle, daemon=True)
    t.start()
    return t


# ════════════════════════════════════════════════════════════════
# CONSULTAS (las usa servidor.py para la API)
# ════════════════════════════════════════════════════════════════
def _filas(cur):
    return [dict(r) for r in cur.fetchall()]


def consultar_averias(q=None, maq=None, mecanico=None, desde=None, hasta=None,
                      tipo=None, estado=None, limite=300):
    sql = """SELECT a.* FROM averias a WHERE 1=1"""
    args = []
    if q:
        # Cada palabra debe aparecer (en cualquier orden) en algún campo de texto
        for palabra in q.split():
            sql += """ AND (a.descripcion LIKE ? OR a.accion LIKE ? OR a.observaciones LIKE ?
                       OR a.parte LIKE ? OR a.maq_nombre LIKE ?)"""
            args += [f'%{palabra}%'] * 5
    if maq:
        sql += ' AND (a.maq_id = ? OR a.maq_nombre LIKE ?)'
        args += [maq, f'%{maq}%']
    if mecanico:
        sql += ' AND a.id IN (SELECT averia_id FROM averia_mecanicos WHERE mecanico LIKE ?)'
        args.append(f'%{mecanico}%')
    if desde:  sql += ' AND a.fecha >= ?'; args.append(desde)
    if hasta:  sql += ' AND a.fecha <= ?'; args.append(hasta)
    if tipo:   sql += ' AND a.tipo = ?';   args.append(tipo)
    if estado: sql += ' AND a.estado = ?'; args.append(estado)
    sql += ' ORDER BY a.fecha DESC, a.ts DESC LIMIT ?'
    args.append(int(limite))

    with conectar() as con:
        avs = _filas(con.execute(sql, args))
        for a in avs:
            a['mecanicos'] = [r['mecanico'] for r in con.execute(
                'SELECT mecanico FROM averia_mecanicos WHERE averia_id=?', (a['id'],))]
            a['materiales'] = _filas(con.execute(
                'SELECT referencia, descripcion, cantidad FROM averia_materiales WHERE averia_id=?',
                (a['id'],)))
            a.pop('json', None)
    return avs


def consultar_maquinas(q=None):
    sql = """
        SELECT m.*,
               (SELECT COUNT(*) FROM averias a WHERE a.maq_id = m.id)      AS n_averias,
               (SELECT MAX(a.fecha) FROM averias a WHERE a.maq_id = m.id)  AS ultima_averia,
               (SELECT e.estado FROM estado_maquinas e WHERE e.maq_id = m.id) AS estado_actual
        FROM maquinas m WHERE 1=1"""
    args = []
    if q:
        sql += ' AND (m.id LIKE ? OR m.nombre LIKE ? OR m.tipo LIKE ? OR m.area LIKE ?)'
        args += [f'%{q}%'] * 4
    sql += ' ORDER BY m.id'
    with conectar() as con:
        return _filas(con.execute(sql, args))


def consultar_repuestos(q=None, bajo_stock=False, limite=300):
    sql = 'SELECT * FROM repuestos WHERE 1=1'
    args = []
    if q:
        # Cada palabra debe aparecer (en cualquier orden) en ref/descripción/ubicación
        for palabra in q.split():
            sql += """ AND (referencia LIKE ? OR descripcion LIKE ? OR ubicacion LIKE ?)"""
            args += [f'%{palabra}%'] * 3
    if bajo_stock:
        sql += ' AND stock IS NOT NULL AND stock_minimo IS NOT NULL AND stock < stock_minimo'
    sql += ' ORDER BY descripcion LIMIT ?'
    args.append(int(limite))
    with conectar() as con:
        return _filas(con.execute(sql, args))


def consultar_kpis():
    with conectar() as con:
        mes = datetime.date.today().isoformat()[:7]
        k = {}
        k['total_averias']    = con.execute('SELECT COUNT(*) c FROM averias').fetchone()['c']
        k['averias_mes']      = con.execute(
            "SELECT COUNT(*) c FROM averias WHERE fecha LIKE ?", (mes+'%',)).fetchone()['c']
        k['averias_abiertas'] = con.execute(
            "SELECT COUNT(*) c FROM averias WHERE estado IS NOT NULL AND estado <> 'RESUELTA'"
        ).fetchone()['c']
        k['tiempo_medio_min'] = con.execute(
            'SELECT ROUND(AVG(tiempo_min),1) v FROM averias WHERE tiempo_min > 0'
        ).fetchone()['v']
        k['por_tipo'] = _filas(con.execute(
            'SELECT tipo, COUNT(*) n FROM averias WHERE tipo IS NOT NULL GROUP BY tipo ORDER BY n DESC'))
        k['top_maquinas'] = _filas(con.execute("""
            SELECT maq_nombre, maq_id, COUNT(*) n, SUM(COALESCE(tiempo_min,0)) min_total
            FROM averias GROUP BY maq_id, maq_nombre ORDER BY n DESC LIMIT 10"""))
        k['top_mecanicos'] = _filas(con.execute("""
            SELECT mecanico, COUNT(*) n FROM averia_mecanicos GROUP BY mecanico ORDER BY n DESC LIMIT 10"""))
        k['materiales_top'] = _filas(con.execute("""
            SELECT referencia, descripcion, SUM(COALESCE(cantidad,0)) total
            FROM averia_materiales GROUP BY referencia, descripcion ORDER BY total DESC LIMIT 10"""))
        k['repuestos_bajo_stock'] = con.execute("""
            SELECT COUNT(*) c FROM repuestos
            WHERE stock IS NOT NULL AND stock_minimo IS NOT NULL
              AND CAST(stock AS REAL) < CAST(stock_minimo AS REAL)
              AND CAST(stock_minimo AS REAL) > 0""").fetchone()['c']
        k['n_maquinas']  = con.execute('SELECT COUNT(*) c FROM maquinas').fetchone()['c']
        k['n_repuestos'] = con.execute('SELECT COUNT(*) c FROM repuestos').fetchone()['c']
        return k


def info_bd():
    with conectar() as con:
        crear_esquema()
        tablas = [r['name'] for r in con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]
        info = {'archivo': str(DB_FILE), 'tablas': {}}
        for t in tablas:
            info['tablas'][t] = con.execute(f'SELECT COUNT(*) c FROM {t}').fetchone()['c']
        ult = con.execute('SELECT * FROM sync_log ORDER BY ts DESC LIMIT 1').fetchone()
        info['ultimo_sync'] = dict(ult) if ult else None
        return info


# ════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print('Sincronizando panalcor.db ...')
    r = sync_completo()
    print('  Excel   :', ', '.join(r['excel']) or '—')
    print('  Firebase:', ', '.join(r['firebase']) or '—')
    i = info_bd()
    print(f"\nBase de datos: {i['archivo']}")
    for t, n in sorted(i['tablas'].items()):
        if n: print(f'  {t:26s} {n:>6} registros')
    print('\nListo.')
